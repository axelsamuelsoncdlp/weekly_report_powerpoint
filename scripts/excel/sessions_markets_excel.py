import os
import pandas as pd
import openpyxl
import platform
import subprocess

def update_sessions_markets_excel():
    """Updates the Excel sheet 'sessions_markets' with finalized sessions data and opens it."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "macros", "top_table.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "sessions_markets_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Detect separator used in CSV ("," or ";")
    with open(CSV_FILE, "r", encoding="utf-8") as file:
        first_line = file.readline()
    sep = ";" if ";" in first_line else ","

    # ✅ Load the data from CSV
    df_raw = pd.read_csv(CSV_FILE, sep=sep, decimal=",", dtype=str)

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_raw.columns.tolist())

    # ✅ Ensure required columns exist
    expected_cols = {"Market", "Year Type"}
    if not expected_cols.issubset(df_raw.columns):
        raise ValueError(f"❌ Missing expected columns! Found: {df_raw.columns}")

    # ✅ Rename "Year Type" to "Year" for consistency with Excel
    df_raw.rename(columns={"Year Type": "Year"}, inplace=True)

    # ✅ Extract ISO Week column names (All numeric columns)
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]
    iso_week_cols.sort(key=int)  # Sort numerically

    print(f"\n📊 **ISO Week Columns Found:** {iso_week_cols}")

    # ✅ Convert numeric columns to float for proper handling
    for col in iso_week_cols:
        df_raw[col] = pd.to_numeric(df_raw[col], errors="coerce")

    # ✅ Create a clean DataFrame with only the columns we need
    df_parsed = df_raw[["Market", "Year"] + iso_week_cols].copy()

    # ✅ Debug: Print the DataFrame before writing to Excel
    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_parsed.to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb["sessions_markets"]

    # ✅ Define starting positions dynamically
    header_row, start_row, start_col = 5, 6, 2

    # ✅ Clear previous data (Dynamic range)
    max_rows = start_row + len(df_parsed)
    max_cols = start_col + len(iso_week_cols) + 2  # +2 for "Market" and "Year"
    
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (including ISO Week columns)
    headers = ["Market", "Year"] + iso_week_cols
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Ensure we write all rows correctly
    for row_idx, row_vals in enumerate(df_parsed.itertuples(index=False), start=start_row):
        market_name = row_vals[0]  # First column in each row is the Market name
        year_name = row_vals[1]  # Second column in each row is the Year

        print(f"\n📝 Writing row {row_idx} → Market: '{market_name}' | Year: '{year_name}'")

        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = headers[col_idx - start_col]  # Get column name (Market, Year, ISO week)

            print(f"   🔹 Column: {col_name} | Raw Value: {val}", end="")

            if isinstance(val, (int, float)):
                if pd.notna(val):
                    cell.value = float(val)  # ✅ Convert to float for proper Excel formatting
                    print(f" → Converted to float: {cell.value}")
                else:
                    cell.value = None  # ✅ Handle NaN values
                    print(f" → Set to None (NaN)")
            else:
                cell.value = val  # ✅ Keep non-numeric values unchanged
                print(f" → Kept as text: '{cell.value}'")


    # ✅ Force Excel refresh and save
    wb.save(EXCEL_FILE)
    
    # ✅ Force recalculation to ensure all formulas/files update
    wb.close()
    
    print("\n✅ Excel file saved and updated with **ALL** sessions markets data!")

    # ✅ Close any existing Excel windows and reopen file
    print("\n📂 **Closing existing Excel windows and opening updated file...**")
    
    # ✅ Kill any existing Excel processes to ensure clean update
    try:
        if platform.system() == "Darwin":  # macOS
            os.system("pkill -f 'Microsoft Excel' 2>/dev/null || true")
            os.system("sleep 2")  # Wait a moment
            os.system(f"open '{EXCEL_FILE}'")
        elif platform.system() == "Windows":  # Windows
            os.system("taskkill /F /IM EXCEL.EXE 2>nul || echo Excel not running")
            subprocess.run(["start", "/wait", "excel", EXCEL_FILE], shell=True)
        elif platform.system() == "Linux":  # Linux
            os.system("pkill -f excel 2>/dev/null || true")
            os.system(f"xdg-open '{EXCEL_FILE}'")
    except Exception as e:
        print(f"⚠️ Warning: Could not close existing Excel: {e}")
        # Fallback: try to open anyway
        if platform.system() == "Darwin":
            os.system(f"open '{EXCEL_FILE}'")
        elif platform.system() == "Windows":
            subprocess.run(["start", "excel", EXCEL_FILE], shell=True)
        elif platform.system() == "Linux":
            os.system(f"xdg-open '{EXCEL_FILE}'")

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_sessions_markets_excel()

