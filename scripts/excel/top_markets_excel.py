import os
import sys
import pandas as pd
import openpyxl

# Add scripts directory to path
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))
from calculator.date_utils import get_latest_full_week

# 📂 Define file paths
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
CSV_FINAL_PATH = os.path.join(BASE_DIR, "data", "final", "top_markets_final.csv")  # Formatted Revenue
CSV_GROWTH_PATH = os.path.join(BASE_DIR, "data", "final", "top_markets_growth_final.csv")  # Revenue Growth
CSV_SHARE_PATH = os.path.join(BASE_DIR, "data", "final", "top_markets_share_final.csv")  # Revenue Share
EXCEL_FILE_PATH = os.path.join(BASE_DIR, "macros", "top_table.xlsm")  # Output Excel file
SHEET_NAME = "top_markets"  # Excel sheet to update

def update_excel_with_top_markets():
    """Updates the Excel sheet with revenue, growth, and revenue share data."""

    # ✅ Load CSV files
    df_revenue = pd.read_csv(CSV_FINAL_PATH)
    df_growth = pd.read_csv(CSV_GROWTH_PATH)
    df_share = pd.read_csv(CSV_SHARE_PATH)

    # ✅ Standardize column names (strip spaces to avoid errors)
    df_revenue.columns = df_revenue.columns.str.strip()
    df_growth.columns = df_growth.columns.str.strip()
    df_share.columns = df_share.columns.str.strip()

    # ✅ Print available columns for debugging
    print("\n🔍 Debug: Columns in df_revenue:", list(df_revenue.columns))
    print("\n🔍 Debug: Columns in df_growth:", list(df_growth.columns))
    print("\n🔍 Debug: Columns in df_share:", list(df_share.columns))

    # ✅ Validate required columns
    required_columns = ["8-week avg"]
    for df, name in zip([df_revenue, df_growth, df_share], ["Revenue", "Growth", "Share"]):
        if "8-week avg" not in df.columns:
            print(f"❌ Error: '8-week avg' column not found in {name} data!")
            return  # Exit function if column is missing

    # ✅ Extract week numbers from column headers (ignore 'Market' and '8-week avg')
    week_columns = [col for col in df_revenue.columns if col not in ["Market", "8-week avg"]]
    week_numbers = week_columns.copy()  # Keep week numbers consistent

    # ✅ Open Excel workbook
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
    ws = wb[SHEET_NAME]

    # 🎯 Update current week date in C2
    latest_week = get_latest_full_week()["current_week"]
    start_date, end_date = latest_week

    def format_date(dt):
        """Formats a date as 'Jan 27th' with proper suffix."""
        suffix = "th" if 11 <= dt.day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(dt.day % 10, "th")
        return dt.strftime(f"%b {dt.day}{suffix}")

    formatted_week = f"{format_date(start_date)} - {format_date(end_date)}"
    ws["C2"] = formatted_week  # ✅ Update the current week cell

    # 🎯 Update week numbers in row 3 (C3:J3, L3:T3, U3:AC3)
    for col_idx, week in enumerate(week_numbers, start=3):  # Week labels in C3-J3, L3-T3, U3-AC3
        ws.cell(row=3, column=col_idx, value=int(week))  # Revenue
        ws.cell(row=3, column=col_idx+9, value=int(week))  # Growth
        ws.cell(row=3, column=col_idx+18, value=int(week))  # Share

    # 🎯 Update market names in column B (B4:B20)
    for row_idx, market in enumerate(df_revenue["Market"], start=4):
        ws.cell(row=row_idx, column=2, value=market)  # Column B

    # 🎯 Insert Gross Revenue data into C4:J20
    start_col_revenue = 3  # Column C
    for row_idx, (_, row) in enumerate(df_revenue.iterrows(), start=4):
        for col_idx, value in enumerate(row[1:9], start=start_col_revenue):  # First 8 weeks (C-J)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 🎯 Insert 8-week avg. into Column K (K4:K20)
    for row_idx, avg_value in enumerate(df_revenue["8-week avg"], start=4):
        ws.cell(row=row_idx, column=11, value=avg_value)  # Column K

    # 🎯 Insert Revenue Growth data into L4:T20 (WITHOUT Market Column)
    start_col_growth = 12  # Column L
    for row_idx, (_, row) in enumerate(df_growth.iterrows(), start=4):
        for col_idx, value in enumerate(row, start=start_col_growth):  # All weeks (L-T)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 🎯 Insert Revenue Share data into U4:AC18
    start_col_share = 21  # Column U
    for row_idx, (_, row) in enumerate(df_share.iterrows(), start=4):
        for col_idx, value in enumerate(row, start=start_col_share):  # All weeks (U-AC)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ✅ Save the updated Excel file
    wb.save(EXCEL_FILE_PATH)

    print("\n✅ **Excel updated successfully with Revenue, Growth, and Revenue Share Data!**")
    print(f"📄 {CSV_FINAL_PATH} → C4:K20 (Revenue & 8-Week Avg.)")
    print(f"📄 {CSV_GROWTH_PATH} → L4:T20 (Growth & 8-Week Avg.)")
    print(f"📄 {CSV_SHARE_PATH} → U4:AC18 (Revenue Share)")

if __name__ == "__main__":
    update_excel_with_top_markets()
