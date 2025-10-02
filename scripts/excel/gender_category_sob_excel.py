import os
import pandas as pd
import openpyxl
import platform
import subprocess
import time


def is_excel_open():
    """Checks if Excel is currently running."""
    if platform.system() == "Windows":
        tasklist = os.popen('tasklist').read().lower()
        return "excel.exe" in tasklist
    elif platform.system() == "Darwin":  # macOS
        process_list = os.popen("ps aux").read()
        return "Microsoft Excel" in process_list
    return False


def close_excel():
    """Closes Excel if it's running."""
    print("\n🛑 **Closing Excel to ensure the latest version is used...**")
    if platform.system() == "Windows":
        os.system("taskkill /F /IM excel.exe")
    elif platform.system() == "Darwin":  # macOS
        os.system("pkill -f 'Microsoft Excel'")
    time.sleep(2)  # Ensure Excel fully closes before proceeding


def open_excel(file_path):
    """Opens the updated Excel file."""
    print("\n📂 **Opening updated Excel file...**")
    if platform.system() == "Darwin":  # macOS
        subprocess.Popen(["open", file_path])
    elif platform.system() == "Windows":
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")


def update_gender_category_sob_excel():
    """Inserts the finalized gender-category share data into the Excel sheet dynamically with full logging."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "gender_category_share_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Check if Excel is open and close it before updating
    excel_was_open = is_excel_open()
    if excel_was_open:
        close_excel()

    # ✅ Load the data from CSV
    df_share = pd.read_csv(CSV_FILE, dtype=str)  # Keep values as strings

    print("\n🔍 **CSV File Successfully Loaded!**")
    print(f"📄 CSV Shape: {df_share.shape}")
    print(f"🔢 Column Names in CSV: {df_share.columns.tolist()}")

    # ✅ Extract column headers (ISO Week numbers and '8-Week Avg')
    week_columns = [col for col in df_share.columns if col.isdigit() or col == "8-Week Avg"]

    print(f"📅 **Extracted Week Columns:** {week_columns}")

    # ✅ Convert values to formatted strings with "%"
    for col in week_columns:
        df_share[col] = df_share[col].apply(lambda x: "-" if pd.isna(x) or x in ["0", "0%"] else f"{x}%" if "%" not in str(x) else x)

    # ✅ Debug: Print the first few rows before writing to Excel
    print("\n📊 **Preview of Data Before Writing to Excel:**")
    print(df_share.head(10).to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)

    ws_name = "gender_category"
    if ws_name not in wb.sheetnames:
        raise ValueError(f"❌ Worksheet '{ws_name}' not found in {EXCEL_FILE}!")

    ws = wb[ws_name]

    # ✅ Define starting positions dynamically
    header_row = 5  # Headers in row 5
    start_row = 6  # Data starts from row 6
    start_col = 22  # Column V (22 in Excel)
    
    # ✅ Find where Grand Total row is in Excel
    grand_total_row = None
    for row in range(1, 30):
        cell_value = ws.cell(row=row, column=2).value
        if cell_value and 'Grand Total' in str(cell_value):
            grand_total_row = row
            break
    
    if grand_total_row:
        print(f"📌 Found Grand Total at row {grand_total_row}")
        # ✅ Adjust data writing to stop before Grand Total row
        max_data_rows = grand_total_row - start_row
        print(f"📌 Will write {max_data_rows} rows of data (rows {start_row} to {start_row + max_data_rows - 1})")
    else:
        max_data_rows = len(df_share)
        print(f"⚠️ Grand Total not found, writing all {max_data_rows} rows")

    print("\n📝 **Starting Data Insertion Process...**")
    print(f"📌 Writing to Excel from Row: {start_row}, Column: {start_col} (V5 onwards)")

    # ✅ Separate Grand Total row from other data
    grand_total_data = df_share[df_share['Product Category'] == 'Grand Total']
    df_share_without_grand_total = df_share[df_share['Product Category'] != 'Grand Total']
    
    # ✅ Clear previous data before inserting new data (including Grand Total row)
    for row in range(start_row, start_row + max_data_rows + 1):  # +1 to include Grand Total row
        for col in range(start_col, start_col + len(week_columns)):
            ws.cell(row=row, column=col, value=None)

    print("\n🧹 **Cleared Previous Data in Target Range**")

    # ✅ Insert headers in row 5
    for col_idx, header in enumerate(week_columns, start=start_col):
        print(f"🖊️ Writing Header: {header} → Cell ({header_row}, {col_idx})")
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Verify column alignment by printing CSV headers and corresponding Excel columns
    print("\n🔄 **Column Mapping (CSV → Excel):**")
    for csv_col, excel_col in zip(week_columns, range(start_col, start_col + len(week_columns))):
        print(f"📌 CSV Column '{csv_col}' → Excel Column {excel_col} (Column {chr(64 + excel_col)})")

    # ✅ Insert category data (without Grand Total)
    for r_idx, row in enumerate(df_share_without_grand_total.itertuples(index=False), start=start_row):
        if r_idx >= start_row + max_data_rows:
            print(f"⚠️ Stopping at row {r_idx} to avoid overwriting Grand Total row")
            break
            
        print(f"\n📝 Writing Row {r_idx}: {row[:3]}...")  # Debugging: print first few columns

        # Skip the first two columns (Gender and Product Category) and only write numeric columns
        numeric_vals = row[2:]  # Skip Gender (index 0) and Product Category (index 1)
        
        for c_idx, (col_name, val) in enumerate(zip(week_columns, numeric_vals), start=start_col):
            print(f"   ➤ Writing '{val}' to Cell ({r_idx}, {c_idx})")
            ws.cell(row=r_idx, column=c_idx, value=val)  # ✅ Already formatted as string
    
    # ✅ Write Grand Total row at the end
    if not grand_total_data.empty and grand_total_row:
        print(f"\n📝 Writing Grand Total Row at row {grand_total_row}...")
        grand_total_row_data = grand_total_data.iloc[0]
        numeric_vals = grand_total_row_data.iloc[2:]  # Skip Gender and Product Category
        
        for c_idx, (col_name, val) in enumerate(zip(week_columns, numeric_vals), start=start_col):
            print(f"   ➤ Writing '{val}' to Cell ({grand_total_row}, {c_idx})")
            ws.cell(row=grand_total_row, column=c_idx, value=val)

    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with Gender & Category Share data!")

    # ✅ If Excel was open before, reopen it
    if excel_was_open:
        open_excel(EXCEL_FILE)


# ✅ If running directly, execute function
if __name__ == "__main__":
    update_gender_category_sob_excel()