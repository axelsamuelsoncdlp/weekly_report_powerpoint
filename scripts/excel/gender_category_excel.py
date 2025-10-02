import os
import pandas as pd
import openpyxl
import platform
import subprocess
import time

def close_excel():
    """Closes Excel to ensure the latest version of the file can be updated."""
    print("\n🛑 **Closing Excel to ensure latest version is used...**")
    
    if platform.system() == "Windows":
        os.system("taskkill /F /IM excel.exe")
    elif platform.system() == "Darwin":  # macOS
        while True:
            # Check if Excel is running
            running = os.system("pgrep -f 'Microsoft Excel' > /dev/null") == 0
            if not running:
                break  # Exit loop if Excel is fully closed
            os.system("pkill -f 'Microsoft Excel'")  # Kill Excel
            time.sleep(1)  # Wait before checking again

    time.sleep(2)  # ✅ Ensures all processes are closed before reopening

def open_excel(file_path):
    """Opens the updated Excel file and ensures it launches successfully."""
    print("\n📂 **Opening updated Excel file...**")
    
    if platform.system() == "Darwin":  # macOS
        os.system(f"open '{file_path}'")
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")
    
    time.sleep(2)  # ✅ Ensure Excel has time to fully open before the script exits
    print("\n✅ **Excel file opened successfully!**")

def update_gender_category_excel():
    """Inserts the finalized gender-category revenue data into the Excel sheet without modifying the order."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "gender_category_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Close Excel before updating
    close_excel()

    # ✅ Load the data from CSV **as it is**
    df_raw = pd.read_csv(CSV_FILE)

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_raw.columns.tolist())

    # ✅ Extract ISO Week column names (All numeric columns)
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]

    # ✅ Convert numeric values back to float/int (formatting in thousands)
    for col in iso_week_cols + ["8-week avg"]:
        df_raw[col] = df_raw[col] / 1000  # ✅ Convert revenue values to thousands

    # ✅ Debug: Print the DataFrame before writing to Excel
    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_raw.to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    
    ws_name = "gender_category"
    if ws_name not in wb.sheetnames:
        raise ValueError(f"❌ Worksheet '{ws_name}' not found in {EXCEL_FILE}!")

    ws = wb[ws_name]

    # ✅ Define starting positions dynamically
    header_row, start_row, start_col = 5, 6, 2  # Adjust based on actual layout

    # ✅ Clear previous data dynamically
    max_rows = start_row + len(df_raw)
    max_cols = start_col + len(iso_week_cols) + 2  # +2 for "Gender" and "8-week avg"
    
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (including ISO Week columns)
    headers = df_raw.columns.tolist()
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Write data to Excel **exactly as it appears in the CSV**
    for row_idx, row_vals in enumerate(df_raw.itertuples(index=False), start=start_row):
        print(f"\n📝 Writing row {row_idx} → {row_vals}")

        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = headers[col_idx - start_col]

            if isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = "#,##0"  # ✅ Keep number formatting in thousands
            else:
                cell.value = val

    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with Gender & Category Revenue data!")

    # ✅ Open Excel after update
    open_excel(EXCEL_FILE)

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_gender_category_excel()