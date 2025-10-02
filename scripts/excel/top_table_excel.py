import os
import sys
import pandas as pd
import openpyxl
import platform
import subprocess
import time
from datetime import datetime

# Ensure correct import paths
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from calculator.date_utils import get_latest_full_week

def is_excel_open():
    """Checks if Excel is currently running."""
    if platform.system() == "Windows":
        tasklist = os.popen('tasklist').read().lower()
        return "excel.exe" in tasklist
    elif platform.system() == "Darwin":  # macOS
        process_list = os.popen("ps aux").read()
        return "Microsoft Excel" in process_list
    return False

def close_excel():
    """Closes Excel if it's running."""
    print("\n🛑 **Closing Excel to ensure the latest version is used...**")
    if platform.system() == "Windows":
        os.system("taskkill /F /IM excel.exe")
    elif platform.system() == "Darwin":  # macOS
        os.system("pkill -f 'Microsoft Excel'")
    time.sleep(3)  # ✅ Ensure Excel fully closes before proceeding

def open_excel(file_path):
    """Opens the updated Excel file."""
    print("\n📂 **Opening updated Excel file...**")
    if platform.system() == "Darwin":  # macOS
        subprocess.Popen(["open", file_path])
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")

def update_top_table():
    """Updates the 'top_table' tab with finalized metrics, growth, and YTD data."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    FINAL_DATA_DIR = os.path.join(BASE_DIR, "data/final")

    # ✅ File paths for finalized data
    data_files = {
        "metrics": os.path.join(FINAL_DATA_DIR, "metrics_final.csv"),
        "growth": os.path.join(FINAL_DATA_DIR, "growth_metrics_final.csv"),
        "ytd_metrics": os.path.join(FINAL_DATA_DIR, "ytd_metrics_final.csv"),
        "ytd_growth": os.path.join(FINAL_DATA_DIR, "ytd_growth_final.csv"),
    }

    # ✅ Ensure all files exist
    for name, file_path in data_files.items():
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"❌ Required data file is missing: {file_path}")

    # ✅ Load data
    dataframes = {name: pd.read_csv(file, index_col=0) for name, file in data_files.items()}

    # ✅ Ensure files are not empty
    for name, df in dataframes.items():
        if df.empty:
            raise ValueError(f"❌ {name} dataset is empty!")

    # ✅ Check if Excel is open and close it before updating
    close_excel()

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    
    # ✅ Ensure the 'top_table' sheet exists
    sheet_name = "top_table"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"❌ Sheet '{sheet_name}' not found in {EXCEL_FILE}!")

    ws = wb[sheet_name]  # ✅ Select the correct sheet

    # **Step 1️⃣: Format and Write the Current Week Date in C2**
    latest_week = get_latest_full_week()["current_week"]
    start_date, end_date = latest_week

    def format_date(dt):
        """Formats a date as 'Jan 27th' with proper suffix."""
        suffix = "th" if 11 <= dt.day <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(dt.day % 10, "th")
        return dt.strftime(f"%b {dt.day}{suffix}")

    formatted_week = f"{format_date(start_date)} - {format_date(end_date)}"
    ws["C2"] = formatted_week  # ✅ Update the current week cell

    # **Step 2️⃣: Extract and Write Month Name to L4 (Fix for Merged Cell Issue)**
    first_day_month = start_date.strftime("%B")

    # ✅ Ensure we only write to the first cell in the merged range
    for merged_range in ws.merged_cells.ranges:
        if "L4" in str(merged_range):
            ws.unmerge_cells(str(merged_range))  # ✅ Unmerge first
            break  # ✅ Exit after unmerging

    ws["L4"] = first_day_month  # ✅ Now writing is possible

    # **Step 3️⃣: Identify Start Positions for Data**
    start_row = 5  # Data starts here
    col_mapping = {
        "metrics": 3,  # Column C
        "growth": 8,   # Column H
        "ytd_metrics": 12,  # Column L
        "ytd_growth": 16,   # Column P
    }

    # ✅ **Clear Previous Data Before Writing**
    max_rows = start_row + len(dataframes["metrics"])
    max_cols = max(col_mapping.values()) + len(dataframes["metrics"].columns)

    for row in range(start_row, max_rows):
        for col in range(3, max_cols + 1):
            ws.cell(row=row, column=col, value=None)

    # **Step 4️⃣: Write Data to Excel**
    for name, df in dataframes.items():
        col_start = col_mapping[name]
        for row_idx, metric in enumerate(df.index, start=start_row):
            values = df.loc[metric].values
            for col_offset, value in enumerate(values):
                ws.cell(row=row_idx, column=col_start + col_offset, value=value)

    # **Step 5️⃣: Adjust Row Heights**
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30  # Uniform height

    # ✅ Save the Excel file
    try:
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"\n✅ Excel file {EXCEL_FILE} has been updated successfully!")
    except Exception as e:
        print(f"\n⚠️ Error saving Excel file: {e}")

    open_excel(EXCEL_FILE)

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_top_table()