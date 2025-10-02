import os
import pandas as pd
import openpyxl
import platform
import subprocess

def update_online_kpis_excel():
    """Updates the Excel sheet 'online_kpis' with finalized KPI data and opens it."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "online_kpis_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Detect separator used in CSV ("," or ";")
    with open(CSV_FILE, "r", encoding="utf-8") as file:
        first_line = file.readline()
    sep = ";" if ";" in first_line else ","

    # ✅ Load the data from CSV
    df_raw = pd.read_csv(CSV_FILE, sep=sep, decimal=",", dtype=str)

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_raw.columns.tolist())

    # ✅ Ensure required columns exist
    expected_cols = {"Metric", "Year Type"}
    if not expected_cols.issubset(df_raw.columns):
        raise ValueError(f"❌ Missing expected columns! Found: {df_raw.columns}")

    # ✅ Rename "Year Type" to "Year" for consistency with Excel
    df_raw.rename(columns={"Year Type": "Year"}, inplace=True)

    # ✅ Extract ISO Week column names (All numeric columns)
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]
    
    # ✅ Convert all numeric values back to float/int
    df_parsed = df_raw.copy()
    for col in iso_week_cols:
        df_parsed[col] = df_parsed[col].str.replace(",", ".")  # Ensure correct float parsing
        df_parsed[col] = pd.to_numeric(df_parsed[col], errors="coerce")  # Convert to number

    # ✅ Ensure "Last Year" appears first
    df_parsed["Year"] = pd.Categorical(df_parsed["Year"], categories=["Last Year", "Current Year"], ordered=True)
    df_parsed = df_parsed.sort_values(by=["Metric", "Year"], ascending=[True, True])

    # ✅ Debug: Check if all metrics are included
    required_metrics = [
        "AOV New Customers (ex. VAT)",
        "AOV Returning Customers (ex. VAT)",
        "COS%",
        "Conversion Rate (%)",
        "New Customers",
        "Online Media Spend",
        "Returning Customers",
        "Sessions",
        "nCAC"
    ]

    missing_metrics = [metric for metric in required_metrics if metric not in df_parsed["Metric"].unique()]
    if missing_metrics:
        print(f"❌ Missing metrics in data: {missing_metrics}")
    else:
        print("✅ All required metrics are included.")

    # ✅ Debug: Print the DataFrame before writing to Excel
    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_parsed.to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb["online_kpis"]

    # ✅ Define starting positions dynamically
    header_row, start_row, start_col = 5, 6, 2

    # ✅ Clear previous data (Dynamic range)
    max_rows = start_row + len(df_parsed)
    max_cols = start_col + len(iso_week_cols) + 2  # +2 for "Metric" and "Year"
    
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (including ISO Week columns)
    headers = ["Metric", "Year"] + iso_week_cols
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Ensure we write all rows correctly
    for row_idx, row_vals in enumerate(df_parsed.itertuples(index=False), start=start_row):
        metric_name = row_vals[0]  # First column in each row is the Metric name
        year_name = row_vals[1]  # Second column in each row is the Year

        print(f"\n📝 Writing row {row_idx} → Metric: '{metric_name}' | Year: '{year_name}'")

        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = headers[col_idx - start_col]  # Get column name (Metric, Year, ISO week)

            print(f"   🔹 Column: {col_name} | Raw Value: {val}", end="")

            if isinstance(val, (int, float)):
                if "Conversion Rate" in metric_name:  
                    cell.value = val / 100  # ✅ Fix percentage scaling
                    cell.number_format = "0.0%"  # ✅ 1 decimal for Conversion Rate
                    print(f" → Formatted as Percentage (1 decimal): {cell.value}")

                elif "COS%" in metric_name:  
                    cell.value = val / 100  # ✅ Fix percentage scaling
                    cell.number_format = "0%"  # ✅ 0 decimals for COS%
                    print(f" → Formatted as Percentage (0 decimals): {cell.value}")

                elif "Sessions" in metric_name:
                    cell.value = val  # ✅ Keep value as is
                    cell.number_format = "#,##0.0"  # ✅ 1 decimal for Sessions
                    print(f" → Formatted as Number (1 decimal): {cell.value}")

                else:
                    cell.value = val  # ✅ Directly write value as is
                    cell.number_format = "#,##0"  # ✅ Standard number formatting
                    print(f" → Written as Number: {cell.value}")

            else:
                cell.value = val  # ✅ Keep non-numeric values unchanged
                print(f" → Kept as text: '{cell.value}'")


    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with **ALL** KPI data!")

    # ✅ Automatically open the Excel file after saving
    print("\n📂 **Opening Excel file...**")
    
    if platform.system() == "Darwin":  # macOS
        os.system(f"open '{EXCEL_FILE}'")
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", EXCEL_FILE], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{EXCEL_FILE}'")

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_online_kpis_excel()