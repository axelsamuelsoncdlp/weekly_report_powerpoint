import os
import pandas as pd
import openpyxl
import platform
import subprocess

def update_conversion_markets_excel():
    """Updates the Excel sheet 'conversion_markets' with finalized conversion rate data and opens it."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "conversion_markets_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    if not os.path.exists(CSV_FILE):
        raise FileNotFoundError(f"❌ CSV file not found: {CSV_FILE}")

    # ✅ Load CSV with proper separator detection
    with open(CSV_FILE, "r", encoding="utf-8") as file:
        first_line = file.readline()
    sep = ";" if ";" in first_line else ","

    df_raw = pd.read_csv(CSV_FILE, sep=";", decimal=",", dtype=str)

    print("\n🔍 **Loaded Column Names:**")
    print(df_raw.columns.tolist())

    # ✅ Validate required columns
    expected_cols = {"Market", "Year Type"}
    if not expected_cols.issubset(df_raw.columns):
        raise ValueError(f"❌ Missing expected columns! Found: {df_raw.columns}")

    # ✅ Rename Year Type to Year for consistency
    df_raw.rename(columns={"Year Type": "Year"}, inplace=True)

    # ✅ Get ISO week columns (numeric columns)
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]
    iso_week_cols.sort(key=int)

    print(f"\n📊 **ISO Week Columns Found:** {iso_week_cols}")

    # ✅ Parse numeric columns (convert % strings to numeric values)
    for col in iso_week_cols:
        df_raw[col] = df_raw[col].apply(lambda x: float(x.replace('%', '')) if isinstance(x, str) and '%' in x else x)

    df_parsed = df_raw[["Market", "Year"] + iso_week_cols].copy()

    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_parsed.to_string(index=False))

    # ✅ Load Excel workbook
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb["conversion_markets"]

    # ✅ Clear existing data but preserve headers/formatting
    header_row, start_row, start_col = 5, 6, 2

    max_rows = start_row + len(df_parsed)
    max_cols = start_col + len(iso_week_cols) + 2

    # Clear data area
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers
    headers = ["Market", "Year"] + iso_week_cols
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Write data rows
    for row_idx, row_vals in enumerate(df_parsed.itertuples(index=False), start=start_row):
        market_name = row_vals[0]
        year_name = row_vals[1]

        print(f"\n📝 Writing row {row_idx} → Market: '{market_name}' | Year: '{year_name}'")

        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = headers[col_idx - start_col]

            print(f"   🔹 Column: {col_name} | Raw Value: {val}", end="")

            if isinstance(val, (int, float)):
                if pd.notna(val):
                    # ✅ Convert percentage values to decimal format for Excel (2.51% → 0.0251)
                    decimal_value = val / 100
                    cell.value = decimal_value
                    
                    # ✅ Apply percentage formatting: 2 decimal places for conversion rates
                    cell.number_format = "0.00%"
                    
                    print(f" → Converted to decimal: {decimal_value:.4f} (will display as {val:.2f}%)")
                else:
                    cell.value = None
                    print(f" → Set to None (NaN)")
            else:
                cell.value = val
                print(f" → Kept as text: '{cell.value}'")

    # ✅ Save Excel file
    wb.save(EXCEL_FILE)
    
    # ✅ Close workbook to release file locks
    wb.close()
    
    print("\n✅ Excel file saved and updated with **ALL** conversion markets data!")

    # ✅ Close any existing Excel windows and reopen file for fresh view
    print("\n📂 **Closing existing Excel windows and opening updated file...**")
    
    try:
        if platform.system() == "Darwin":  # macOS
            os.system("pkill -f 'Microsoft Excel' 2>/dev/null || true")
            os.system("sleep 2")  # Wait a moment for process to terminate
            os.system(f"open '{EXCEL_FILE}'")
        elif platform.system() == "Windows":  # Windows
            os.system("system(\"taskkill /F /IM EXCEL.EXE 2>nul || echo Excel not running\")")
            subprocess.run(["start", "/wait", "excel", EXCEL_FILE], shell=True)
        elif platform.system() == "Linux":  # Linux
            os.system("pkill -f excel 2>/dev/null || true")
            os.system(f"xdg-open '{EXCEL_FILE}'")
        else:
            print("⚠️ Unknown platform, cannot auto-open Excel")
    except Exception as e:
        print(f"⚠️ Warning: Could not close existing Excel: {e}")
        # Fallback: try to open anyway
        try:
            if platform.system() == "Darwin":
                os.system(f"open '{EXCEL_FILE}'")
            elif platform.system() == "resource://org.mozilla.uriloader/":
                subprocess.run(["start", "excel", EXCEL_FILE], shell=True)
            elif platform.system() == "Linux":
                os.system(f"xdg-open '{EXCEL_FILE}'")
        except:
            print(f"⚠️ Could not open Excel automatically. Please manually open: {EXCEL_FILE}")

if __name__ == "__main__":
    update_conversion_markets_excel()
