
import os
import pandas as pd
import openpyxl
import platform
import subprocess
import time

def close_excel():
    """Closes Excel to ensure the latest version of the file can be updated."""
    print("\n🛑 **Closing Excel to ensure latest version is used...**")
    if platform.system() == "Windows":
        os.system("taskkill /F /IM excel.exe")
    elif platform.system() == "Darwin":  # macOS
        os.system("pkill -f 'Microsoft Excel'")
    time.sleep(2)  # ✅ Ensure Excel fully closes before proceeding

def open_excel(file_path):
    """Opens the updated Excel file."""
    print("\n📂 **Opening updated Excel file...**")
    if platform.system() == "Darwin":  # macOS
        subprocess.Popen(["open", file_path])
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")

def update_gender_category_growth_excel():
    """Inserts the finalized gender-category growth data into the Excel sheet at column 13, row 5."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "gender_category_growth_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Close Excel before updating
    close_excel()

    # ✅ Load the data from CSV
    df_growth = pd.read_csv(CSV_FILE)

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_growth.columns.tolist())

    # ✅ Extract ISO Week column names (all numeric)
    week_columns = [col for col in df_growth.columns if col.isdigit()]

    # ✅ Convert values to integers (rounding to remove decimals)
    for col in week_columns + ["8-week avg"]:
        df_growth[col] = df_growth[col].round().astype("Int64")  # Keeps NaN as <NA> instead of converting to float

    # ✅ Debug: Print the DataFrame before writing to Excel
    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_growth.to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    
    ws_name = "gender_category"
    if ws_name not in wb.sheetnames:
        raise ValueError(f"❌ Worksheet '{ws_name}' not found in {EXCEL_FILE}!")

    ws = wb[ws_name]

    # ✅ Define starting positions (column 13 = M, row 5)
    header_row, start_row, start_col = 5, 6, 13  # Adjusted for correct placement

    # ✅ Clear previous data dynamically
    max_rows = start_row + len(df_growth)
    max_cols = start_col + len(week_columns) + 1  # +1 for "8-week avg"

    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (ISO Week columns)
    headers = week_columns + ["8-week avg"]
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Write data to Excel **exactly as it appears in the CSV**
    for row_idx, row_vals in enumerate(df_growth.itertuples(index=False), start=start_row):
        print(f"\n📝 Writing row {row_idx} → {row_vals}")

        # Skip the first column (Product Category) and only write numeric columns
        numeric_vals = row_vals[1:]  # Skip the first column (Product Category)
        
        for col_idx, val in enumerate(numeric_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if col_idx - start_col < len(headers):
                col_name = headers[col_idx - start_col]
                
                if pd.notna(val) and val != 0 and val != "<NA>":  # ✅ If the value is not NaN, 0, or "<NA>"
                    try:
                        val_num = int(val)  # Convert to int to remove decimals
                        if val_num < 0:
                            cell.value = f"({abs(val_num)}%)"  # ✅ Format negatives with parentheses
                        else:
                            cell.value = f"{val_num}%"  # ✅ Keep positive numbers with %
                    except (ValueError, TypeError):
                        cell.value = "-"  # ✅ Replace non-numeric values with "-"
                else:
                    cell.value = "-"  # ✅ Replace 0, NaN, or "<NA>" with "-"

    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with Gender & Category Growth data!")

    # ✅ Open Excel after update
    open_excel(EXCEL_FILE)

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_gender_category_growth_excel()