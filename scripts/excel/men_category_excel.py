import os
import pandas as pd
import openpyxl
import platform
import subprocess
import time

def close_excel():
    """Attempts to close any open instances of Excel before updating the file."""
    if platform.system() == "Windows":
        os.system("taskkill /IM excel.exe /F")
    elif platform.system() == "Darwin":  # macOS
        os.system("pkill -f 'Microsoft Excel'")
    elif platform.system() == "Linux":
        os.system("pkill -f 'libreoffice'")
    
    time.sleep(2)  # Wait a moment to ensure Excel has fully closed


def open_excel(file_path):
    """Opens the Excel file after updating to ensure the user sees the latest version."""
    print("\n📂 **Opening Excel file...**")
    
    if platform.system() == "Darwin":  # macOS
        os.system(f"open '{file_path}'")
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")


def update_men_category_excel():
    """Updates the Excel sheet 'men_category' with finalized category revenue data for men, formatting numbers in thousands."""

    # ✅ Close Excel before updating
    print("\n🛑 Closing Excel to ensure file is not locked...")
    close_excel()

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "macros", "top_table.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "men_category_revenue_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Detect separator used in CSV ("," or ";")
    with open(CSV_FILE, "r", encoding="utf-8") as file:
        first_line = file.readline()
    sep = ";" if ";" in first_line else ","

    # ✅ Load the data from CSV
    df_raw = pd.read_csv(CSV_FILE, sep=sep, decimal=",", dtype=str)

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_raw.columns.tolist())

    # ✅ Ensure required columns exist
    expected_cols = {"Gender", "Product Category", "Year Type"}
    if not expected_cols.issubset(df_raw.columns):
        raise ValueError(f"❌ Missing expected columns! Found: {df_raw.columns}")

    # ✅ Rename columns for consistency with Excel
    df_raw.rename(columns={"Year Type": "Year", "Product Category": "Category"}, inplace=True)

    # ✅ Extract ISO Week column names (All numeric columns)
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]
    
    # ✅ Convert all numeric values back to float/int and format in thousands
    df_parsed = df_raw.copy()
    for col in iso_week_cols:
        df_parsed[col] = df_parsed[col].str.replace(",", ".")  # Ensure correct float parsing
        df_parsed[col] = pd.to_numeric(df_parsed[col], errors="coerce")  # Convert to number
        df_parsed[col] = df_parsed[col] / 1000  # ✅ Convert revenue values to thousands

    # ✅ Ensure "Last Year" appears first
    df_parsed["Year"] = pd.Categorical(df_parsed["Year"], categories=["Last Year", "Current Year"], ordered=True)
    df_parsed = df_parsed.sort_values(by=["Gender", "Category", "Year"], ascending=[True, True, True])

    # ✅ Debug: Check if all categories are included
    missing_categories = df_parsed["Category"].isnull().sum()
    if missing_categories > 0:
        print(f"❌ Missing category values in data: {missing_categories} rows have null categories!")
    else:
        print("✅ All required category data is included.")

    # ✅ Debug: Print the DataFrame before writing to Excel
    print("\n📊 **Final Data (Before Writing to Excel):**")
    print(df_parsed.to_string(index=False))

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    
    ws_name = "men_category"
    if ws_name not in wb.sheetnames:
        raise ValueError(f"❌ Worksheet '{ws_name}' not found in {EXCEL_FILE}!")

    ws = wb[ws_name]

    # ✅ Define starting positions dynamically
    header_row, start_row, start_col = 5, 6, 2  # Adjust based on actual layout

    # ✅ Clear previous data dynamically
    max_rows = start_row + len(df_parsed)
    max_cols = start_col + len(iso_week_cols) + 3  # +3 for "Gender", "Category", and "Year"
    
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (including ISO Week columns)
    headers = ["Gender", "Category", "Year"] + iso_week_cols
    print("\n📝 **Writing Headers to Excel:**", headers)
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Write data to Excel
    for row_idx, row_vals in enumerate(df_parsed.itertuples(index=False), start=start_row):
        gender = row_vals[0]  # First column is Gender
        category = row_vals[1]  # Second column is Category
        year_name = row_vals[2]  # Third column is Year

        print(f"\n📝 Writing row {row_idx} → Gender: '{gender}' | Category: '{category}' | Year: '{year_name}'")

        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = headers[col_idx - start_col]  # Get column name

            print(f"   🔹 Column: {col_name} | Raw Value: {val}", end="")

            if isinstance(val, (int, float)):
                cell.value = val  # ✅ Keep values as is (already in thousands)
                cell.number_format = "#,##0"  # ✅ Standard number formatting (in thousands)
                print(f" → Written as Number (in thousands): {cell.value}")
            else:
                cell.value = val  # ✅ Keep non-numeric values unchanged
                print(f" → Kept as text: '{cell.value}'")


    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with Men's Category Revenue data (formatted in thousands)!")

    # ✅ Open Excel after updating
    open_excel(EXCEL_FILE)

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_men_category_excel()