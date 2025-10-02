import os
import time
import pandas as pd
import openpyxl
import platform
import subprocess

def close_excel():
    """Stänger alla Excel-processer för att säkerställa att filen kan uppdateras."""
    if platform.system() == "Windows":
        os.system("taskkill /f /im excel.exe")  # Windows: Stäng Excel
    elif platform.system() == "Darwin":
        os.system("pkill -f 'Microsoft Excel'")  # macOS: Stäng Excel
    elif platform.system() == "Linux":
        os.system("pkill -f 'libreoffice'")  # Linux: Stäng LibreOffice
    print("✅ Stängde alla Excel-fönster.")

def open_excel(file_path):
    """Öppnar Excel-filen efter att den uppdaterats."""
    print("\n📂 **Öppnar Excel-fil...**")
    if platform.system() == "Darwin":
        os.system(f"open '{file_path}'")
    elif platform.system() == "Windows":
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":
        os.system(f"xdg-open '{file_path}'")

def update_gender_excel():
    """Uppdaterar Excel-filen och säkerställer att den öppnas korrekt efter uppdatering."""

    # ✅ Definiera filvägar
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "data", "weekly_report.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data", "final", "gender_revenue_final.csv")

    # ✅ Stäng Excel innan vi börjar
    close_excel()
    time.sleep(3)  # Vänta några sekunder för att säkerställa att Excel-processen har avslutats

    # ✅ Kontrollera att filen finns
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ Filen hittades inte: {EXCEL_FILE}")

    # ✅ Hämta CSV-data
    with open(CSV_FILE, "r", encoding="utf-8") as file:
        first_line = file.readline()
    sep = ";" if ";" in first_line else ","

    df_raw = pd.read_csv(CSV_FILE, sep=sep, decimal=",", dtype=str)

    # ✅ Säkerställ att nödvändiga kolumner finns
    expected_cols = {"Metric", "Year Type"}
    if not expected_cols.issubset(df_raw.columns):
        raise ValueError(f"❌ Saknade kolumner! Finns: {df_raw.columns}")

    # ✅ Byt namn på kolumnen
    df_raw.rename(columns={"Year Type": "Year"}, inplace=True)

    # ✅ Extrahera veckokolumner
    iso_week_cols = [col for col in df_raw.columns if col.isdigit()]
    
    # ✅ Konvertera numeriska värden
    df_parsed = df_raw.copy()
    for col in iso_week_cols:
        df_parsed[col] = df_parsed[col].str.replace(",", ".")
        df_parsed[col] = pd.to_numeric(df_parsed[col], errors="coerce")

    # ✅ Sortera "Last Year" före "Current Year"
    df_parsed["Year"] = pd.Categorical(df_parsed["Year"], categories=["Last Year", "Current Year"], ordered=True)
    df_parsed = df_parsed.sort_values(by=["Metric", "Year"], ascending=[True, True])

    # ✅ Kontrollera att alla förväntade metrics finns
    required_metrics = ["Gross Revenue - MEN", "Gross Revenue - WOMEN"]
    missing_metrics = [metric for metric in required_metrics if metric not in df_parsed["Metric"].unique()]
    
    if missing_metrics:
        print(f"❌ Saknade metrics: {missing_metrics}")
    else:
        print("✅ Alla nödvändiga metrics finns med.")

    # ✅ Öppna och uppdatera Excel-filen
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    ws = wb["gender"]

    # ✅ Definiera positioner
    header_row, start_row, start_col = 5, 6, 2

    # ✅ Rensa gammal data
    max_rows = start_row + len(df_parsed)
    max_cols = start_col + len(iso_week_cols) + 2  # +2 för "Metric" och "Year"
    
    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Skriv headers
    headers = ["Metric", "Year"] + iso_week_cols
    for col_idx, header in enumerate(headers, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Skriv data till Excel
    for row_idx, row_vals in enumerate(df_parsed.itertuples(index=False), start=start_row):
        for col_idx, val in enumerate(row_vals, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx)

            if isinstance(val, (int, float)):
                cell.value = val / 1000  # ✅ Konvertera till tusental
                cell.number_format = "#,##0"  # ✅ Tusentalsformat
            else:
                cell.value = val  # ✅ Behåll text

    # ✅ Spara och stäng filen
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel uppdaterad med senaste gender revenue-data!")

    # ✅ Öppna Excel igen
    open_excel(EXCEL_FILE)

# ✅ Kör funktionen om filen körs direkt
if __name__ == "__main__":
    update_gender_excel()