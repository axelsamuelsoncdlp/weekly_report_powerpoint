import os
import pandas as pd
import openpyxl
import platform
import subprocess
import time

def is_excel_open():
    """Checks if Excel is currently running."""
    if platform.system() == "Windows":
        tasklist = os.popen('tasklist').read().lower()
        return "excel.exe" in tasklist
    elif platform.system() == "Darwin":  # macOS
        process_list = os.popen("ps aux").read()
        return "Microsoft Excel" in process_list
    return False

def close_excel():
    """Closes Excel if it's running."""
    print("\n🛑 **Closing Excel to ensure the latest version is used...**")
    if platform.system() == "Windows":
        os.system("taskkill /F /IM excel.exe")
    elif platform.system() == "Darwin":  # macOS
        os.system("pkill -f 'Microsoft Excel'")
    time.sleep(3)  # ✅ Ensure Excel fully closes before proceeding

def open_excel(file_path):
    """Opens the updated Excel file."""
    print("\n📂 **Opening updated Excel file...**")
    if platform.system() == "Darwin":  # macOS
        subprocess.Popen(["open", file_path])
    elif platform.system() == "Windows":  # Windows
        subprocess.run(["start", "excel", file_path], shell=True)
    elif platform.system() == "Linux":  # Linux (requires LibreOffice or Excel equivalent)
        os.system(f"xdg-open '{file_path}'")

def update_products_women_excel():
    """Inserts the finalized top 20 women's products into the Excel sheet with formatted values."""

    # ✅ Define file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    EXCEL_FILE = os.path.join(BASE_DIR, "macros", "top_table.xlsm")
    CSV_FILE = os.path.join(BASE_DIR, "data/final/products_women_final.csv")

    # ✅ Ensure the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"❌ File not found: {EXCEL_FILE}")

    # ✅ Ensure the CSV file exists
    if not os.path.exists(CSV_FILE):
        raise FileNotFoundError(f"❌ File not found: {CSV_FILE}")

    # ✅ Check if Excel is open and close it before updating
    excel_was_open = is_excel_open()
    if excel_was_open:
        close_excel()

    # ✅ Load the data from CSV
    df_products = pd.read_csv(CSV_FILE, dtype=str)  # Keep data as strings

    # ✅ Convert numeric columns properly
    df_products["Gross Revenue"] = pd.to_numeric(df_products["Gross Revenue"], errors="coerce")
    df_products["Sales Qty"] = pd.to_numeric(df_products["Sales Qty"], errors="coerce")
    df_products["SOB%"] = pd.to_numeric(df_products["SOB%"], errors="coerce")

    # ✅ Ensure proper calculation of SOB% using "Grand Total"
    grand_total_revenue = df_products[df_products["Rank"] == "Grand Total"]["Gross Revenue"].values

    if len(grand_total_revenue) > 0 and pd.notna(grand_total_revenue[0]) and grand_total_revenue[0] > 0:
        df_products["SOB%"] = df_products["Gross Revenue"] / grand_total_revenue[0] * 100
    else:
        df_products["SOB%"] = 0  # Avoid division by zero

    # ✅ Format values:
    df_products["Gross Revenue"] = df_products["Gross Revenue"].apply(lambda x: f"{x / 1000:.1f}" if pd.notna(x) else "-")  # Convert to thousands with 1 decimal
    df_products["SOB%"] = df_products["SOB%"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "-")

    # ✅ Debug: Print loaded column names
    print("\n🔍 **Loaded Column Names:**")
    print(df_products.columns.tolist())

    # ✅ Define starting positions:
    header_row = 5  # Headers in row 5
    start_row = 6  # Data starts from row 6 (below header)
    start_col = 5  # Column V (22 in Excel)

    # ✅ Open the Excel file and select the correct sheet
    wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=True)
    
    ws_name = "products_women"
    if ws_name not in wb.sheetnames:
        raise ValueError(f"❌ Worksheet '{ws_name}' not found in {EXCEL_FILE}!")

    ws = wb[ws_name]

    # ✅ Clear previous data dynamically
    max_rows = start_row + len(df_products)
    max_cols = start_col + len(df_products.columns)

    for row in range(start_row, max_rows):
        for col in range(start_col, max_cols):
            ws.cell(row=row, column=col, value=None)

    # ✅ Write headers to row 5 (column names)
    print("\n📝 **Writing Headers to Excel:**", df_products.columns.tolist())
    for col_idx, header in enumerate(df_products.columns, start=start_col):
        ws.cell(row=header_row, column=col_idx, value=header)

    # ✅ Write entire DataFrame to Excel
    for row_idx, row_vals in enumerate(df_products.itertuples(index=False), start=start_row):
        print(f"\n📝 Writing row {row_idx} → {row_vals[:3]}...")  # Debugging: print first few columns

        for col_idx, val in enumerate(row_vals, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # ✅ Save and close the Excel file
    wb.save(EXCEL_FILE)
    wb.close()

    print("\n✅ Excel successfully updated with Top 20 Women's Products!")

    # ✅ If Excel was open before, reopen it
    if excel_was_open:
        open_excel(EXCEL_FILE)

# ✅ If running directly, execute function
if __name__ == "__main__":
    update_products_women_excel()